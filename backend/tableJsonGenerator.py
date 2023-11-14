from openpyxl import load_workbook
from heatmapGenerator import get_heatmap
import json

def get_table_json(filepath, del_row, graph, curr_state):
    # Load the workbook
    wb = load_workbook(filename=filepath)
    # Get the first sheet
    ws = wb.worksheets[0]
    
    # Remove empty row (generated when df->elsx)
    ws.delete_rows(del_row)

    # Get all rows
    all_rows = list(ws.iter_rows(min_row=1,
                                max_row=ws.max_row,
                                min_col=1,
                                max_col=ws.max_column,
                                values_only=True))


    row_info = []
    col_info = []
    col_info_tree = []
    col_mode = True
    split_index = [0]
    row_depth = 0
    col_depth = 0
    row_index = -1


    for k in range(len(all_rows)):
        table_row = all_rows[k]
        table_row_len = len(table_row)
        cell_info = {}
        # Process each row
        if col_mode:
            col_depth += 1
            is_empty_cell = True
            col_index = 0
            for i in range(table_row_len):
                cell = table_row[i]
                if is_empty_cell:
                    if cell is not None:
                        is_empty_cell = False
                        cell_info['name'] = cell
                        cell_info['children'] = [None]
                        cell_info['colSpan'] = 0
                        col_info.append(cell_info)
                    elif k == 0:
                        row_depth += 1
                else:
                    col_index += 1
                    if cell is not None:
                        end_index = col_index - 1
                        start_index = end_index - cell_info['colSpan']
                        cell_info['colSpan'] = [start_index, end_index]

                        cell_info = {}
                        cell_info['name'] = cell
                        cell_info['children'] = [None]
                        cell_info['colSpan'] = 0
                        col_info.append(cell_info)
                    else:
                        cell_info['children'].append(None)
                        cell_info['colSpan'] += 1
            end_index = col_index
            start_index = end_index - cell_info['colSpan']
            cell_info['colSpan'] = [start_index, end_index]
            split_index.append(len(col_info))

            if all_rows[k+1][0] is not None:
                col_mode = False
                # for i in range(split_index[-2], split_index[-1]):
                #     col_info[i]['children'] = None

                # addChildren2Parent([])
                # for i in range(len(split_index) - 2, -1, -2):
                #     parent_index = split_index[i-1]
                #     child_index = split_index[i]
                #     child_index_now = child_index
                #     for j in range(parent_index, child_index):
                #         children = col_info[j]['children']
                #         children_len = len(children)
                #         for k in range(children_len):
                #             children[k] = col_info[child_index_now]
                #             child_index_now += 1
                # for i in range(split_index[1]):
                #     col_info_tree.append(col_info[i])

        else:

            focus_cell = None
            is_new_row = False

            row_index += 1
            for i in range(row_depth):
                cell = table_row[i]
                cell_info = {}

                if cell is not None:
                    cell_info['name'] = cell
                    cell_info['children'] = []
                    cell_info['rowSpan'] = 0

                    if i == row_depth-1:
                        cell_info['children'] = None
                        cell_info['rowSpan'] = [row_index, row_index]
                        if i != 0:
                            focus_cell['children'].append(cell_info)
                            focus_cell = cell_info
                        else:
                            is_new_row = True
                            row_info.append(cell_info)
                            focus_cell = cell_info

                    else:
                        if i == 0:
                            is_new_row = True
                            if row_index != 0:
                                lastSibling = row_info[-1]
                                addChildren2LastSibling(lastSibling)
                            row_info.append(cell_info)
                            focus_cell = cell_info
                        else:

                            if is_new_row:
                                focus_cell['children'].append(cell_info)
                                focus_cell = cell_info

                            else:

                                if (len(focus_cell['children']) > 0):
                                    lastSibling = focus_cell['children'][-1]
                                    addChildren2LastSibling(lastSibling)
                                focus_cell['children'].append(cell_info)
                                focus_cell = cell_info

                else:
                    if i == 0:
                        focus_cell = row_info[-1]

                    else:
                        focus_cell = focus_cell['children'][-1]

    # print(row_info)
    addChildren2LastSibling(row_info[-1])


    for i in range(split_index[-2], split_index[-1]):
        col_info[i]['children'] = None
    # 构造 col tree
    for i in range(len(split_index) - 1, 1, -1):
        subset = split_index[i-2:i+1]
        # print(subset)
        addChildren2Parent(col_info, [subset[0], subset[1]], [
            subset[1], subset[2]])
    for i in range(split_index[0], split_index[1]):
        col_info_tree.append(col_info[i])

    # with open('table.json', 'w') as f:
    #     json.dump({
    #         'table':
    #         {
    #             'colDepth': col_depth,
    #             'rowDepth': row_depth,
    #             'rowNum': row_index+1,
    #             'colNum': col_index+1,
    #             'rows': row_info,
    #             'cols': col_info_tree
    #         }

    #     }, f

    structure = {
        'colDepth': col_depth,
        'rowDepth': row_depth,
        'rowNum': row_index+1,
        'colNum': col_index+1,
        'rows': row_info,
        'cols': col_info_tree
    }

    heat = get_heatmap(structure, graph, curr_state)

    table = {'structure': structure, 'heat': heat} 

    return table

def addChildren2Parent(colInfo, parentRange, childrenRange):
    childIndex = childrenRange[0]-1
    for i in range(parentRange[0], parentRange[1]):
        parentInfo = colInfo[i]
        parentSpanLen = parentInfo['colSpan'][1] - parentInfo['colSpan'][0]+1
        children = []
        childrenSpanLen = 0
        while (childrenSpanLen < parentSpanLen):
            childIndex += 1
            childInfo = colInfo[childIndex]
            childSpanLen = childInfo['colSpan'][1] - childInfo['colSpan'][0]+1
            childrenSpanLen += childSpanLen
            children.append(childInfo)
            # print(childrenSpanLen,parentSpanLen)
        if (len(children) > 0):
            # print(children)
            parentInfo['children'] = children
        else:
            parentInfo['children'] = None


def addChildren2LastSibling(lastSibling):

    # lastSiblingLastChild = lastSibling['children'][-1]
    lastSiblingLastChildren = lastSibling['children']
    if (lastSiblingLastChildren):
        if isinstance(lastSiblingLastChildren[-1]['rowSpan'], int):
            addChildren2LastSibling(lastSiblingLastChildren[-1])
        lastSibling['rowSpan'] = [
            lastSibling['children'][0]['rowSpan'][0], lastSibling['children'][-1]['rowSpan'][1]]